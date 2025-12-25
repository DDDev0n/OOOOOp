from models import Computer, Processor, RAM, Motherboard
from memento import Caretaker
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def main():
    print("=" * 50)
    print("ОБЪЕКТНО-ОРИЕНТИРОВАННОЕ ПРОЕКТИРОВАНИЕ")
    print("Система управления компонентами компьютера")
    print("=" * 50)

    # Создаем компьютер
    my_pc = Computer("Gaming PC 2025")
    caretaker = Caretaker()

    # Добавляем компоненты (шаг 1)
    print("\n--- Этап 1: Добавляем компоненты ---")
    my_pc.add_component(Processor("Intel Core i9-14900K", "Intel", cores=24, frequency_ghz=6.0))
    my_pc.add_component(Motherboard("ASUS ROG Maximus Z890", "ASUS", socket="LGA1700", max_ram_gb=192))
    my_pc.add_component(RAM("Kingston Fury Beast", "Kingston", capacity_gb=32, speed_mhz=6000))

    # Сохраняем состояние (Memento)
    caretaker.save({"components": [c.get_info() for c in my_pc.components]})
    my_pc.list_components()

    # Добавляем еще компоненты (шаг 2)
    print("\n--- Этап 2: Добавляем дополнительные компоненты ---")
    my_pc.add_component(RAM("Kingston Fury Beast", "Kingston", capacity_gb=32, speed_mhz=6000))
    caretaker.save({"components": [c.get_info() for c in my_pc.components]})
    my_pc.list_components()

    # История изменений
    caretaker.list_history()

    # Диаграмма классов в текстовом формате
    print("\n" + "=" * 50)
    print("ДИАГРАММА КЛАССОВ")
    print("=" * 50)
    print("""
    ┌─────────────────┐
    │     Detail      │  (Абстрактный класс)
    ├─────────────────┤
    │ - name: str     │
    │ - type: Enum    │
    │ - specs: dict   │
    └────────┬────────┘
             │
    ┌────────┴────────────────────────┐
    │        Наследование             │
    │                                 │
    ▼                ▼                ▼
┌──────────┐  ┌──────────┐  ┌───────────────┐
│Processor │  │   RAM    │  │ Motherboard   │
└──────────┘  └──────────┘  └───────────────┘

    ┌─────────────────┐
    │   Computer      │  (Композиция)
    ├─────────────────┤
    │ - components[]  │◄──── Detail[]
    └─────────────────┘
    """)

    # Экспорт в Excel
    print("\n--- Экспорт в Excel ---")
    export_to_excel(my_pc)


def export_to_excel(computer: Computer):
    """Экспорт информации о компонентах в Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Компоненты"

    # Стили
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки
    headers = ["№", "Название", "Тип", "Производитель", "Характеристики"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Данные компонентов
    for idx, component in enumerate(computer.components, 1):
        specs = ", ".join([f"{k}: {v}" for k, v in component.specifications.items()])

        row_data = [
            idx,
            component.name,
            component.component_type.value,
            component.manufacturer,
            specs
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=idx + 1, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Автоширина колонок
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40

    filename = "components_report.xlsx"
    wb.save(filename)
    print(f"✓ Файл {filename} создан успешно!")

    print("\n✓ Программа завершена успешно!")


if __name__ == "__main__":
    main()
